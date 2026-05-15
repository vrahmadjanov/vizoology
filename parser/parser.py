from __future__ import annotations

from collections.abc import Iterator

from django.conf import settings
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string


def parse_excel_column_letter(letter: str) -> int:
    """
    Буквенное обозначение колонки Excel (A, B, ..., Z, AA, ...) в 1-based индекс openpyxl.
    """
    normalized = letter.strip().upper()
    if not normalized:
        raise ValueError("Буква колонки не может быть пустой.")
    try:
        return column_index_from_string(normalized)
    except ValueError as exc:
        raise ValueError(f"Некорректная буква колонки: {letter!r}") from exc


def _question_text_from_cell(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def iter_nonempty_questions_in_column(
    worksheet: Worksheet,
    question_column_letter: str | None = None,
    *,
    start_row: int = 1,
    end_row: int | None = None,
) -> Iterator[tuple[int, str]]:
    """
    Обходит лист построчно, начиная с start_row, без особой трактовки строки заголовка.

    Колонка вопросов по умолчанию — settings.PARSER_DEFAULT_QUESTION_COLUMN_LETTER
    (переменная окружения PARSER_DEFAULT_QUESTION_COLUMN_LETTER).
    Если передан непустой question_column_letter, используется он.
    Ячейки с пустым значением после strip пропускаются.
    end_row включительно; по умолчанию — worksheet.max_row.
    """
    if start_row < 1:
        raise ValueError("start_row должен быть >= 1.")
    letter = (
        question_column_letter
        if question_column_letter is not None
        else settings.PARSER_DEFAULT_QUESTION_COLUMN_LETTER
    )
    col_idx = parse_excel_column_letter(letter)
    last = end_row if end_row is not None else worksheet.max_row
    if last < start_row:
        return
    for row in range(start_row, last + 1):
        raw = worksheet.cell(row=row, column=col_idx).value
        text = _question_text_from_cell(raw)
        if text is not None:
            yield row, text


def resolve_first_answer_column_index(
    *,
    question_column_letter: str | None = None,
    answer_block_start_column_letter: str | None = None,
) -> int:
    """
    Индекс первой колонки блока «ответ | источники | рассуждения» (1-based).

    Если answer_block_start_column_letter задан — используется он.
    Иначе берётся колонка сразу после колонки вопросов.
    Колонка вопросов по умолчанию — settings.PARSER_DEFAULT_QUESTION_COLUMN_LETTER.
    """
    q_letter = (
        question_column_letter
        if question_column_letter is not None
        else settings.PARSER_DEFAULT_QUESTION_COLUMN_LETTER
    )
    question_idx = parse_excel_column_letter(q_letter)
    if answer_block_start_column_letter is not None:
        return parse_excel_column_letter(answer_block_start_column_letter)
    return question_idx + 1


def write_three_column_answer_block(
    worksheet: Worksheet,
    row: int,
    *,
    first_answer_column_index: int,
    answer_text: str,
    sources_text: str,
    reasoning_text: str,
) -> None:
    """Пишет три ячейки подряд: ответ, источники, рассуждения."""
    worksheet.cell(row=row, column=first_answer_column_index).value = answer_text
    worksheet.cell(
        row=row, column=first_answer_column_index + 1
    ).value = sources_text
    worksheet.cell(
        row=row, column=first_answer_column_index + 2
    ).value = reasoning_text

