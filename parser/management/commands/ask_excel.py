from __future__ import annotations

from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import DatabaseError
from openpyxl import load_workbook

from ai.management.commands.ask import (
    _save_history,
    _sources_for_answer,
    _unique_sources,
)
from ai.rag import RAGAnswer, answer_question
from parser.parser import (
    iter_nonempty_questions_in_column,
    resolve_first_answer_column_index,
    write_three_column_answer_block,
)


class Command(BaseCommand):
    help = (
        "Читает вопросы из колонки Excel, для каждой строки вызывает RAG и записывает "
        "ответ в ту же книгу (три колонки: ответ, источники, рассуждения), затем сохраняет файл."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "workbook",
            type=str,
            help="Путь к файлу .xlsx (результат пишется в этот же файл).",
        )
        parser.add_argument(
            "--sheet",
            type=str,
            default="",
            help="Имя листа; по умолчанию активный лист.",
        )
        parser.add_argument(
            "--questions-col",
            type=str,
            default="",
            help=(
                "Буква колонки с вопросами (например Q). "
                "Пусто — из PARSER_DEFAULT_QUESTION_COLUMN_LETTER в settings."
            ),
        )
        parser.add_argument(
            "--answers-start-col",
            type=str,
            default="",
            help=(
                "Буква первой колонки блока ответа (3 колонки подряд). "
                "Пусто — сразу справа от колонки вопросов."
            ),
        )
        parser.add_argument(
            "--top-k",
            type=int,
            default=5,
            help="Сколько ближайших фрагментов использовать как контекст.",
        )
        parser.add_argument(
            "--min-score",
            type=float,
            default=settings.RAG_MIN_SCORE,
            help="Минимальный score лучшего фрагмента для вызова Gemini.",
        )
        parser.add_argument(
            "--no-history",
            action="store_true",
            help="Не сохранять записи в QuestionAnswerHistory.",
        )

    def handle(self, *args, **options):
        path = Path(options["workbook"]).expanduser().resolve()
        if not path.is_file():
            raise CommandError(f"Файл не найден: {path}")
        if path.suffix.lower() != ".xlsx":
            raise CommandError("Ожидается файл с расширением .xlsx")

        top_k = options["top_k"]
        min_score = options["min_score"]
        if top_k < 1:
            raise CommandError("--top-k должен быть больше 0.")
        if not 0 <= min_score <= 1:
            raise CommandError("--min-score должен быть от 0 до 1.")

        questions_col = options["questions_col"].strip() or None
        answers_start = options["answers_start_col"].strip() or None
        sheet_name = options["sheet"].strip() or None

        try:
            wb = load_workbook(path)
        except Exception as exc:
            raise CommandError(f"Не удалось открыть книгу: {exc}") from exc

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise CommandError(
                    f"Лист «{sheet_name}» не найден. Есть: {', '.join(wb.sheetnames)}"
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active
        if ws is None:
            raise CommandError("В книге нет активного листа.")

        first_ans_col = resolve_first_answer_column_index(
            question_column_letter=questions_col,
            answer_block_start_column_letter=answers_start,
        )

        count = 0
        errors = 0
        for row, question in iter_nonempty_questions_in_column(
            ws, questions_col or None
        ):
            try:
                rag_answer = answer_question(
                    question, top_k=top_k, min_score=min_score
                )
            except Exception as exc:
                errors += 1
                msg = f"Ошибка RAG: {exc}"
                self.stderr.write(self.style.WARNING(f"Строка {row}: {msg}"))
                write_three_column_answer_block(
                    ws,
                    row,
                    first_answer_column_index=first_ans_col,
                    answer_text=msg,
                    sources_text="",
                    reasoning_text="",
                )
                continue

            sources_cell = _rag_sources_to_cell(rag_answer)
            write_three_column_answer_block(
                ws,
                row,
                first_answer_column_index=first_ans_col,
                answer_text=rag_answer.structured_answer.short_answer,
                sources_text=sources_cell,
                reasoning_text=rag_answer.structured_answer.reasoning_summary,
            )
            count += 1
            if not options["no_history"]:
                try:
                    _save_history(rag_answer, top_k=top_k, min_score=min_score)
                except DatabaseError as exc:
                    raise CommandError(
                        "Не удалось сохранить историю в БД (миграции ai?). "
                        f"Подробнее: {exc}"
                    ) from exc
            self.stdout.write(f"Строка {row}: готово")

        try:
            wb.save(path)
        except Exception as exc:
            raise CommandError(f"Не удалось сохранить книгу: {exc}") from exc

        self.stdout.write(
            self.style.SUCCESS(
                f"Сохранено в {path}: обработано вопросов {count}, с ошибками {errors}."
            )
        )


def _rag_sources_to_cell(rag_answer: RAGAnswer) -> str:
    sources_list = _sources_for_answer(rag_answer)
    if not sources_list and rag_answer.sources:
        sources_list = _unique_sources(rag_answer.sources)
    lines: list[str] = []
    for source in sources_list:
        url = source.url or "без ссылки"
        lines.append(
            f"- {source.title}: {url} (chunk_id={source.chunk_id}, score={source.score:.4f})"
        )
    return "\n".join(lines)
