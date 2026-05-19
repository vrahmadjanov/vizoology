from __future__ import annotations

from collections.abc import Iterator, Sequence

from django.conf import settings
from openpyxl.styles import Alignment, Font
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ANSWER_HEADER_LABELS: tuple[str, str, str] = ("Ответ", "Источники", "Рассуждения")
ANSWER_COLUMN_WIDTHS: tuple[float, float, float] = (40, 50, 50)
ANSWER_BLOCK_HEADER_ROW: int = 1
ANSWER_ROW_MIN_HEIGHT: float = 60

_ANSWER_CELL_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
_ANSWER_HEADER_FONT = Font(bold=True)
_SOURCE_LINK_FONT = Font(color="0563C1", underline="single")


def format_source_line(number: int, title: str) -> str:
    """«N. Заголовок» — одна строка в колонке источников."""
    return f"{number}. {title}"


def write_source_cell(
    cell,
    *,
    number: int,
    title: str,
    url: str | None,
) -> None:
    cell.value = format_source_line(number, title)
    _apply_answer_cell_format(cell)
    if url:
        cell.hyperlink = url
        cell.font = _SOURCE_LINK_FONT


def parse_excel_column_letter(letter: str) -> int:
    """
    Буквенное обозначение колонки Excel (A, B, ..., Z, AA, ...) в 1-based индекс openpyxl.
    """
    normalized = letter.strip().upper()
    if not normalized:
        raise ValueError("Буква колонки не может быть пустой.")
    try:
        return column_index_from_string(normalized)
    except ValueError as exc:
        raise ValueError(f"Некорректная буква колонки: {letter!r}") from exc


def _question_text_from_cell(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def iter_nonempty_questions_in_column(
    worksheet: Worksheet,
    question_column_letter: str | None = None,
    *,
    start_row: int = 1,
    end_row: int | None = None,
) -> Iterator[tuple[int, str]]:
    """
    Обходит лист построчно, начиная с start_row, без особой трактовки строки заголовка.

    Колонка вопросов по умолчанию — settings.PARSER_DEFAULT_QUESTION_COLUMN_LETTER
    (переменная окружения PARSER_DEFAULT_QUESTION_COLUMN_LETTER).
    Если передан непустой question_column_letter, используется он.
    Ячейки с пустым значением после strip пропускаются.
    end_row включительно; по умолчанию — worksheet.max_row.
    """
    if start_row < 1:
        raise ValueError("start_row должен быть >= 1.")
    letter = (
        question_column_letter
        if question_column_letter is not None
        else settings.PARSER_DEFAULT_QUESTION_COLUMN_LETTER
    )
    col_idx = parse_excel_column_letter(letter)
    last = end_row if end_row is not None else worksheet.max_row
    if last < start_row:
        return
    for row in range(start_row, last + 1):
        raw = worksheet.cell(row=row, column=col_idx).value
        text = _question_text_from_cell(raw)
        if text is not None:
            yield row, text


def resolve_first_answer_column_index(
    *,
    question_column_letter: str | None = None,
    answer_block_start_column_letter: str | None = None,
) -> int:
    """
    Индекс первой колонки блока «ответ | источники | рассуждения» (1-based).

    Если answer_block_start_column_letter задан — используется он.
    Иначе берётся колонка сразу после колонки вопросов.
    Колонка вопросов по умолчанию — settings.PARSER_DEFAULT_QUESTION_COLUMN_LETTER.
    """
    q_letter = (
        question_column_letter
        if question_column_letter is not None
        else settings.PARSER_DEFAULT_QUESTION_COLUMN_LETTER
    )
    question_idx = parse_excel_column_letter(q_letter)
    if answer_block_start_column_letter is not None:
        return parse_excel_column_letter(answer_block_start_column_letter)
    return question_idx + 1


def _apply_answer_cell_format(cell) -> None:
    cell.alignment = _ANSWER_CELL_ALIGNMENT


def ensure_answer_block_headers(
    worksheet: Worksheet,
    first_answer_column_index: int,
    *,
    header_row: int = ANSWER_BLOCK_HEADER_ROW,
) -> None:
    """Заголовки блока ответа в первой строке (только в пустые ячейки)."""
    for offset, label in enumerate(ANSWER_HEADER_LABELS):
        cell = worksheet.cell(
            row=header_row, column=first_answer_column_index + offset
        )
        if cell.value in (None, ""):
            cell.value = label
        cell.alignment = _ANSWER_CELL_ALIGNMENT
        cell.font = _ANSWER_HEADER_FONT


def apply_answer_block_column_widths(
    worksheet: Worksheet,
    first_answer_column_index: int,
) -> None:
    """Ширина колонок блока ответа (символы Excel)."""
    for offset, width in enumerate(ANSWER_COLUMN_WIDTHS):
        letter = get_column_letter(first_answer_column_index + offset)
        worksheet.column_dimensions[letter].width = width


def apply_answer_row_min_height(
    worksheet: Worksheet,
    row: int,
    *,
    min_height: float = ANSWER_ROW_MIN_HEIGHT,
) -> None:
    """Минимальная высота строки для читаемости многострочного текста."""
    dim = worksheet.row_dimensions[row]
    if dim.height is None or dim.height < min_height:
        dim.height = min_height


def _excel_formula_string(value: str) -> str:
    return value.replace('"', '""')


def sources_cell_formula(sources: Sequence[tuple[int, str, str | None]]) -> str:
    """
    Формула Excel: каждый источник на новой строке, заголовок — кликабельная ссылка.
    """
    if not sources:
        return ""
    parts: list[str] = []
    for number, title, url in sources:
        line = _excel_formula_string(format_source_line(number, title))
        if url:
            safe_url = _excel_formula_string(url)
            parts.append(f'HYPERLINK("{safe_url}","{line}")')
        else:
            parts.append(f'"{line}"')
    if len(parts) == 1:
        return f"={parts[0]}"
    return "=" + "&CHAR(10)&".join(parts)


def write_sources_cell(
    cell,
    sources: Sequence[tuple[int, str, str | None]],
) -> None:
    """Одна ячейка: «N. Заголовок» на каждой строке, заголовок — гиперссылка."""
    if not sources:
        cell.value = ""
        _apply_answer_cell_format(cell)
        return

    if len(sources) == 1:
        number, title, url = sources[0]
        if url:
            write_source_cell(
                cell, number=number, title=title, url=url
            )
            return
        cell.value = format_source_line(number, title)
        _apply_answer_cell_format(cell)
        return

    cell.value = sources_cell_formula(sources)
    _apply_answer_cell_format(cell)
    cell.font = _SOURCE_LINK_FONT


def write_three_column_answer_block(
    worksheet: Worksheet,
    row: int,
    *,
    first_answer_column_index: int,
    answer_text: str,
    sources: Sequence[tuple[int, str, str | None]] | None = None,
    sources_text: str = "",
    reasoning_text: str,
) -> None:
    """Пишет три ячейки подряд: ответ, источники, рассуждения."""
    answer_cell = worksheet.cell(row=row, column=first_answer_column_index)
    answer_cell.value = answer_text
    _apply_answer_cell_format(answer_cell)

    sources_cell = worksheet.cell(row=row, column=first_answer_column_index + 1)
    if sources is not None:
        write_sources_cell(sources_cell, sources)
    else:
        sources_cell.value = sources_text
        _apply_answer_cell_format(sources_cell)

    reasoning_cell = worksheet.cell(
        row=row, column=first_answer_column_index + 2
    )
    reasoning_cell.value = reasoning_text
    _apply_answer_cell_format(reasoning_cell)

    apply_answer_row_min_height(worksheet, row)
