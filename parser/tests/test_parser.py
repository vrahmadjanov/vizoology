import re
import tempfile
from io import BytesIO
from unittest.mock import patch

from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import transaction
from django.test import Client, SimpleTestCase, TestCase, TransactionTestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ai.rag import RAGAnswer, SourceSnippet, StructuredAnswer
from parser.models import ExcelAskJob
from parser.services.excel_batch import (
    _RowFillResult,
    _collect_row_results,
    _excel_ask_max_workers,
    fill_workbook_rag,
    rag_sources_for_column,
)
from parser.services.excel_job_runner import process_excel_job
from parser.services.parser import (
    ANSWER_COLUMN_WIDTHS,
    ANSWER_HEADER_LABELS,
    ANSWER_ROW_MIN_HEIGHT,
    apply_answer_block_column_widths,
    ensure_answer_block_headers,
    format_source_line,
    iter_nonempty_questions_in_column,
    parse_excel_column_letter,
    resolve_first_answer_column_index,
    sources_cell_formula,
    write_three_column_answer_block,
)


class ParseExcelColumnLetterTests(SimpleTestCase):
    def test_known_columns(self) -> None:
        self.assertEqual(parse_excel_column_letter("A"), 1)
        self.assertEqual(parse_excel_column_letter("z"), 26)
        self.assertEqual(parse_excel_column_letter(" AA "), 27)

    def test_empty_raises(self) -> None:
        with self.assertRaisesMessage(ValueError, "Буква колонки не может быть пустой"):
            parse_excel_column_letter("  ")

    def test_invalid_raises(self) -> None:
        with self.assertRaises(ValueError):
            parse_excel_column_letter("1")


class IterNonemptyQuestionsTests(SimpleTestCase):
    @override_settings(PARSER_DEFAULT_QUESTION_COLUMN_LETTER="Q")
    def test_default_question_column_from_settings_q(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["Q1"] = "from_q"
        ws["A1"] = "not_used"
        rows = list(iter_nonempty_questions_in_column(ws))
        self.assertEqual(rows, [(1, "from_q")])

    @override_settings(PARSER_DEFAULT_QUESTION_COLUMN_LETTER="R")
    def test_default_question_column_follows_setting(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["R2"] = "only_r"
        rows = list(iter_nonempty_questions_in_column(ws))
        self.assertEqual(rows, [(2, "only_r")])

    def test_starts_row_one_no_header_semantics_skips_blank(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "first"
        ws["A2"] = None
        ws["A3"] = "  third  "
        ws["A4"] = ""
        ws["A5"] = "last"
        rows = list(iter_nonempty_questions_in_column(ws, "A"))
        self.assertEqual(rows, [(1, "first"), (3, "third"), (5, "last")])

    def test_range_and_save_roundtrip(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["B2"] = "q1"
        ws["B4"] = "q2"
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        from openpyxl import load_workbook

        wb2 = load_workbook(buf)
        ws2 = wb2.active
        assert ws2 is not None
        rows = list(iter_nonempty_questions_in_column(ws2, "B", start_row=1, end_row=10))
        self.assertEqual(rows, [(2, "q1"), (4, "q2")])


class AnswerBlockColumnTests(SimpleTestCase):
    @override_settings(PARSER_DEFAULT_QUESTION_COLUMN_LETTER="Q")
    def test_default_first_answer_after_question(self) -> None:
        self.assertEqual(
            resolve_first_answer_column_index(),
            parse_excel_column_letter("R"),
        )

    def test_explicit_start_overrides_offset(self) -> None:
        self.assertEqual(
            resolve_first_answer_column_index(
                question_column_letter="B",
                answer_block_start_column_letter="E",
            ),
            parse_excel_column_letter("E"),
        )

    def test_write_three_columns(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        write_three_column_answer_block(
            ws,
            2,
            first_answer_column_index=2,
            answer_text="a",
            sources_text="s",
            reasoning_text="r",
        )
        self.assertEqual(ws["B2"].value, "a")
        self.assertEqual(ws["C2"].value, "s")
        self.assertEqual(ws["D2"].value, "r")
        for col in ("B", "C", "D"):
            cell = ws[f"{col}2"]
            self.assertTrue(cell.alignment.wrap_text)
            self.assertEqual(cell.alignment.vertical, "top")
        self.assertEqual(ws.row_dimensions[2].height, ANSWER_ROW_MIN_HEIGHT)

    def test_answer_block_headers_and_column_widths(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ensure_answer_block_headers(ws, first_answer_column_index=2)
        apply_answer_block_column_widths(ws, first_answer_column_index=2)
        self.assertEqual(ws["B1"].value, ANSWER_HEADER_LABELS[0])
        self.assertEqual(ws["C1"].value, ANSWER_HEADER_LABELS[1])
        self.assertEqual(ws["D1"].value, ANSWER_HEADER_LABELS[2])
        self.assertTrue(ws["B1"].font.bold)
        for offset, width in enumerate(ANSWER_COLUMN_WIDTHS):
            letter = get_column_letter(2 + offset)
            self.assertEqual(ws.column_dimensions[letter].width, width)


class FillWorkbookRagTests(SimpleTestCase):
    _sample_answer = RAGAnswer(
        question="ignored",
        structured_answer=StructuredAnswer(
            short_answer="Кратко",
            reasoning_summary="Подробно",
            source_numbers=[1],
        ),
        sources=[
            SourceSnippet(
                number=1,
                title="Страница",
                url="https://wiki.example/page",
                chunk_id=42,
                chunk_position=0,
                score=0.91,
                text="контент",
            )
        ],
        model="stub",
    )

    @patch("parser.services.excel_batch.answer_question", autospec=True)
    def test_fills_three_columns_after_question_without_history(self, mock_aq):
        mock_aq.return_value = self._sample_answer
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A2"] = "Вопрос?"
        stats = fill_workbook_rag(
            wb,
            questions_col="A",
            save_history=False,
            top_k=3,
            min_score=0.5,
        )
        self.assertEqual(stats.processed, 1)
        self.assertEqual(stats.errors, 0)
        self.assertEqual(ws["B1"].value, "Ответ")
        self.assertEqual(ws["C1"].value, "Источники")
        self.assertEqual(ws["D1"].value, "Рассуждения")
        self.assertEqual(ws["B2"].value, "Кратко")
        self.assertEqual(ws["C2"].value, format_source_line(1, "Страница"))
        self.assertEqual(ws["C2"].hyperlink.target, "https://wiki.example/page")
        self.assertEqual(ws["D2"].value, "Подробно")
        self.assertTrue(ws["B2"].alignment.wrap_text)
        self.assertTrue(ws["C2"].alignment.wrap_text)
        self.assertEqual(ws.column_dimensions["B"].width, 40)

    @patch("parser.services.excel_batch.answer_question", autospec=True)
    def test_sources_column_multiple_hyperlinks_in_one_cell(self, mock_aq):
        mock_aq.return_value = RAGAnswer(
            question="ignored",
            structured_answer=StructuredAnswer(
                short_answer="Кратко",
                reasoning_summary="Подробно",
                source_numbers=[1, 2],
            ),
            sources=[
                SourceSnippet(
                    number=1,
                    title="Первая",
                    url="https://wiki.example/one",
                    chunk_id=1,
                    chunk_position=0,
                    score=0.9,
                    text="a",
                ),
                SourceSnippet(
                    number=2,
                    title="Вторая",
                    url="https://wiki.example/two",
                    chunk_id=2,
                    chunk_position=0,
                    score=0.8,
                    text="b",
                ),
            ],
            model="stub",
        )
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A2"] = "Вопрос?"
        fill_workbook_rag(
            wb,
            questions_col="A",
            save_history=False,
            top_k=3,
            min_score=0.5,
        )
        formula = ws["C2"].value
        self.assertIsInstance(formula, str)
        self.assertTrue(formula.startswith("="))
        self.assertIn("HYPERLINK", formula)
        self.assertIn("CHAR(10)", formula)
        self.assertIn("Первая", formula)
        self.assertIn("Вторая", formula)

    def test_rag_sources_for_column_and_formula(self) -> None:
        answer = FillWorkbookRagTests._sample_answer
        rows = rag_sources_for_column(answer)
        self.assertEqual(rows, [(1, "Страница", "https://wiki.example/page")])
        formula = sources_cell_formula(rows * 2 + [(2, "Вторая", "https://x.example/y")])
        self.assertIn("CHAR(10)", formula)
        self.assertEqual(formula.count("HYPERLINK"), 3)

    def test_raises_on_unknown_sheet(self) -> None:
        wb = Workbook()
        with self.assertRaisesMessage(ValueError, "Лист «Missing»"):
            fill_workbook_rag(wb, sheet_name="Missing", save_history=False)

    @override_settings(EXCEL_ASK_MAX_WORKERS=4)
    @patch("parser.services.excel_batch.ThreadPoolExecutor")
    @patch("parser.services.excel_batch.answer_question", autospec=True)
    def test_uses_thread_pool_for_multiple_questions(
        self, mock_aq, mock_executor_cls
    ) -> None:
        mock_aq.return_value = self._sample_answer
        mock_executor_cls.return_value.__enter__.return_value.map.side_effect = (
            lambda func, items: [func(item) for item in items]
        )

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A2"] = "Q1"
        ws["A3"] = "Q2"
        stats = fill_workbook_rag(
            wb,
            questions_col="A",
            save_history=False,
            top_k=3,
            min_score=0.5,
        )
        self.assertEqual(stats.processed, 2)
        mock_executor_cls.assert_called_once_with(max_workers=2)
        self.assertEqual(mock_aq.call_count, 2)
        self.assertEqual(ws["B2"].value, "Кратко")
        self.assertEqual(ws["B3"].value, "Кратко")

    @override_settings(EXCEL_ASK_MAX_WORKERS=8)
    def test_max_workers_capped_by_question_count(self) -> None:
        self.assertEqual(_excel_ask_max_workers(1, max_workers=None), 1)
        self.assertEqual(_excel_ask_max_workers(3, max_workers=None), 3)
        self.assertEqual(_excel_ask_max_workers(20, max_workers=None), 8)

    @patch("parser.services.excel_batch._process_question_row")
    def test_collect_row_results_sequential_when_one_worker(
        self, mock_process
    ) -> None:
        mock_process.return_value = _RowFillResult(
            orig_row=2,
            answer_text="ok",
            reasoning_text="",
            sources=[],
            rag_answer=None,
        )
        _collect_row_results(
            [(2, "q")],
            top_k=3,
            min_score=0.5,
            save_history=False,
            warn_row=None,
            info_row=None,
            max_workers=1,
        )
        mock_process.assert_called_once()


class ExcelAskViewTests(SimpleTestCase):
    URL = "/presentation/ask/"

    def test_get_shows_upload_form(self) -> None:
        response = Client().get(self.URL)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "multipart/form-data")


class ExcelJobHistoryViewTests(TestCase):
    def test_list_empty_200(self) -> None:
        url = reverse("presentation_excel_job_history")
        r = Client().get(url)
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Пока нет ни одного задания.")

    def test_list_shows_job(self) -> None:
        wb = Workbook()
        bio = BytesIO()
        wb.save(bio)
        content = bio.getvalue()
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp):
                job = ExcelAskJob.objects.create(
                    original_filename="batch.xlsx",
                    sheet="Лист1",
                    questions_col="Q",
                    answers_start_col="R",
                    top_k=5,
                    min_score=0.55,
                    save_history=True,
                )
                job.input_file.save("in.xlsx", ContentFile(content))
        url = reverse("presentation_excel_job_history")
        r = Client().get(url)
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "batch.xlsx")
        self.assertContains(r, "Лист1")
        job_url = reverse("presentation_ask_job", kwargs={"pk": str(job.pk)})
        self.assertContains(r, job_url)


class ExcelAskPostTests(TestCase):
    @patch("presentation.services.excel_ask._start_excel_job_thread")
    @patch.object(transaction, "on_commit", side_effect=lambda func: func())
    def test_upload_redirects_to_job_page(self, mock_thread, _mock_on_commit) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Вопрос"
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        upload = SimpleUploadedFile(
            "batch.xlsx",
            bio.read(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        resp = Client().post(
            reverse("presentation_ask"),
            data={
                "sheet": "",
                "questions_col": "A",
                "answers_start_col": "",
                "top_k": "5",
                "min_score": "0.55",
                "save_history": "on",
                "workbook": upload,
            },
        )
        self.assertEqual(resp.status_code, 302)
        m = re.search(
            r"/presentation/ask/jobs/([0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-"
            r"[0-9a-f]{4}-[0-9a-f]{12})/",
            resp["Location"],
        )
        self.assertIsNotNone(m)
        job = ExcelAskJob.objects.get(pk=m.group(1))
        self.assertEqual(job.status, ExcelAskJob.Status.PENDING)
        self.assertEqual(job.original_filename, "batch.xlsx")
        mock_thread.assert_called_once()


class ExcelJobRunnerFileTests(TransactionTestCase):
    @patch("parser.services.excel_batch.answer_question", autospec=True)
    def test_process_excel_job_writes_result(self, mock_aq):
        mock_aq.return_value = FillWorkbookRagTests._sample_answer
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "?"
        bio = BytesIO()
        wb.save(bio)
        content = bio.getvalue()
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp):
                job = ExcelAskJob.objects.create(
                    original_filename="t.xlsx",
                    sheet="",
                    questions_col="A",
                    answers_start_col="",
                    top_k=3,
                    min_score=0.5,
                    save_history=False,
                )
                job.input_file.save("in.xlsx", ContentFile(content))
                process_excel_job(job.pk)
                job.refresh_from_db()
                self.assertEqual(job.status, ExcelAskJob.Status.DONE)
                self.assertTrue(job.result_file.name)


class ExcelAskJobApiTests(TransactionTestCase):
    def test_status_json_returns_pending(self) -> None:
        wb = Workbook()
        bio = BytesIO()
        wb.save(bio)
        content = bio.getvalue()
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp):
                job = ExcelAskJob.objects.create(
                    original_filename="t.xlsx",
                    sheet="",
                    questions_col="A",
                    answers_start_col="",
                    top_k=5,
                    min_score=0.55,
                    save_history=False,
                )
                job.input_file.save("in.xlsx", ContentFile(content))

        r = Client().get(
            reverse(
                "presentation_ask_job_status", kwargs={"pk": str(job.pk)}
            )
        )
        self.assertEqual(r.status_code, 200)
        payload = r.json()
        self.assertEqual(payload["status"], ExcelAskJob.Status.PENDING)
