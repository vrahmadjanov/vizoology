from io import BytesIO

from django.test import SimpleTestCase, override_settings
from openpyxl import Workbook

from parser.parser import (
    iter_nonempty_questions_in_column,
    parse_excel_column_letter,
    resolve_first_answer_column_index,
    write_three_column_answer_block,
)


class ParseExcelColumnLetterTests(SimpleTestCase):
    def test_known_columns(self) -> None:
        self.assertEqual(parse_excel_column_letter("A"), 1)
        self.assertEqual(parse_excel_column_letter("z"), 26)
        self.assertEqual(parse_excel_column_letter(" AA "), 27)

    def test_empty_raises(self) -> None:
        with self.assertRaisesMessage(ValueError, "Буква колонки не может быть пустой"):
            parse_excel_column_letter("  ")

    def test_invalid_raises(self) -> None:
        with self.assertRaises(ValueError):
            parse_excel_column_letter("1")


class IterNonemptyQuestionsTests(SimpleTestCase):
    @override_settings(PARSER_DEFAULT_QUESTION_COLUMN_LETTER="Q")
    def test_default_question_column_from_settings_q(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["Q1"] = "from_q"
        ws["A1"] = "not_used"
        rows = list(iter_nonempty_questions_in_column(ws))
        self.assertEqual(rows, [(1, "from_q")])

    @override_settings(PARSER_DEFAULT_QUESTION_COLUMN_LETTER="R")
    def test_default_question_column_follows_setting(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["R2"] = "only_r"
        rows = list(iter_nonempty_questions_in_column(ws))
        self.assertEqual(rows, [(2, "only_r")])

    def test_starts_row_one_no_header_semantics_skips_blank(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "first"
        ws["A2"] = None
        ws["A3"] = "  third  "
        ws["A4"] = ""
        ws["A5"] = "last"
        rows = list(iter_nonempty_questions_in_column(ws, "A"))
        self.assertEqual(rows, [(1, "first"), (3, "third"), (5, "last")])

    def test_range_and_save_roundtrip(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["B2"] = "q1"
        ws["B4"] = "q2"
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        from openpyxl import load_workbook

        wb2 = load_workbook(buf)
        ws2 = wb2.active
        assert ws2 is not None
        rows = list(iter_nonempty_questions_in_column(ws2, "B", start_row=1, end_row=10))
        self.assertEqual(rows, [(2, "q1"), (4, "q2")])


class AnswerBlockColumnTests(SimpleTestCase):
    @override_settings(PARSER_DEFAULT_QUESTION_COLUMN_LETTER="Q")
    def test_default_first_answer_after_question(self) -> None:
        self.assertEqual(
            resolve_first_answer_column_index(),
            parse_excel_column_letter("R"),
        )

    def test_explicit_start_overrides_offset(self) -> None:
        self.assertEqual(
            resolve_first_answer_column_index(
                question_column_letter="B",
                answer_block_start_column_letter="E",
            ),
            parse_excel_column_letter("E"),
        )

    def test_write_three_columns(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        write_three_column_answer_block(
            ws,
            2,
            first_answer_column_index=2,
            answer_text="a",
            sources_text="s",
            reasoning_text="r",
        )
        self.assertEqual(ws["B2"].value, "a")
        self.assertEqual(ws["C2"].value, "s")
        self.assertEqual(ws["D2"].value, "r")
